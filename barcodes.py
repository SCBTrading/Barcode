import os
from openpyxl import load_workbook
from reportlab.lib.pagesizes import landscape
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.graphics.barcode import code128

def generate_barcode_pdf(barcode_data, output_path):
    c = canvas.Canvas(output_path, pagesize=landscape((60*mm, 30*mm)))
    barcode = code128.Code128(barcode_data, barWidth=0.35*mm, barHeight=15*mm)
    barcode_width = barcode.width
    barcode.drawOn(c, (60*mm - barcode_width) / 2, 9*mm)

    # Add barcode number below the barcode
    c.setFont("Helvetica", 12)
    barcode_number_width = c.stringWidth(barcode_data, "Helvetica", 12)
    c.drawString((60*mm - barcode_number_width) / 2, 3*mm, barcode_data)
    
    c.save()

def main():
    # Load Excel spreadsheet
    workbook = load_workbook('barcodes.xlsx')
    sheet = workbook.active

    # Create output directory if not exists
    output_dir = 'barcode_pdfs'
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Iterate over each cell in the first column (assuming barcodes are in column A)
    for row in sheet.iter_rows(min_row=1, max_row=101, min_col=1, max_col=1, values_only=True):
        for cell_value in row:
            if cell_value is not None:
                barcode_data = str(cell_value).zfill(13)  # Ensure barcode has 13 digits
                output_filename = f'{barcode_data}.pdf'
                output_path = os.path.join(output_dir, output_filename)
                generate_barcode_pdf(barcode_data, output_path)
                print(f'Generated barcode PDF for {barcode_data}.')

    print('All barcodes generated successfully.')

if __name__ == "__main__":
    main()

